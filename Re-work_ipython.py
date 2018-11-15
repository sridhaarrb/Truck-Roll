import openpyxl
import re
import pandas as pd

wb = openpyxl.load_workbook("C:\Users\eubcefm\PycharmProjects\Re-Work\Corrective_Task.xlsx")
ws = wb.active
row_count=0
DL_Filtered=[]

yy = str(input('Enter a year: '))
mm = str(input('Enter a month: '))
ee = input('Enter a day: ')

date1=mm+'/'+str(ee)+'/'+yy
nd=int(ee)+1
ee1=str(ee)

date2=str(yy+'-'+mm+'-'+ee1)
date3=str(yy+'-'+mm+'-'+str(nd))

Date_Conf=raw_input('Running for Last Day of Month?(y/n): ')
if Date_Conf=='y':
    nm=str(int(mm)+1)
    date2=str(yy+'-'+mm+'-'+ee1)
    date3=str(yy+'-'+nm+'-'+'1')
    

#date1=mm+'/'+str(ee)+'/'+yy
#nd=int(ee)+1
#ee1=str(ee)

#date2=str(yy+'-'+mm+'-'+ee1)
#date3=str(yy+'-'+mm+'-'+str(nd))

for h in range(2,2000):
    if (ws.cell(row=h, column=22).value)!=None:
        row_count+=1

print 'Total Tickets to be checked: ',row_count

for i in range(2, row_count+2):
    DL = ws.cell(row=i, column=22).value
    Wo_Ref=ws.cell(row=i, column=1).value
    print 'WO Ref taken from the corrective task excel', Wo_Ref
    Yesterday_Comment=[]
    DL1 = DL.encode("utf-8")
    #print type (DL1)
    #print DL1
    #print len(DL1.splitlines())
    #pattern=re.compile(r'\d+/\d+/\d+\s\d+:\d+:\d+\s[A-Z]{2}')
    pattern=re.compile(r'\d+:\d+:\d+\s[A-Z]{2}')
    DL_Filtered=re.split(pattern, DL1)
    #print DL_Filtered
    for j in DL_Filtered:
        if date1 in j:
            Yesterday_Comment.append(j.lower())
            #print 'The Yesterday_Comment List: ',Yesterday_Comment
    
    #print Yesterday_Comment
    #print 'Length: ',len(Yesterday_Comment)
    
    Yesterday_Comment_Rev=list(Yesterday_Comment)
    Yesterday_Comment_Rev.reverse()
    #Day_DL=' '
    
    #for i in Yesterday_Comment:
     #   Day_DL=Day_DL+str(i)
        
    
    if (len(Yesterday_Comment)==0):
        ws.cell(row=i, column=26).value='No logs on mentioned dates'
        ws.cell(row=i, column=27).value='No logs on mentioned dates'
        ws.cell(row=i, column=28).value='No logs on mentioned dates'
        wb.save('C:\Users\eubcefm\PycharmProjects\Re-Work\Corrective_Task.xlsx')
        continue

########################################
#####    Pandas Section            #####
########################################               
    df=pd.read_excel('C:\Users\eubcefm\PycharmProjects\Re-Work\Test.xlsx')
    if (Wo_Ref not in df['Nocrefid'].values):
        break
    else:
        df2=df[(df['Nocrefid']==Wo_Ref) & (df['Audit Timestamp'] >= date2) & (df['Audit Timestamp'] < date3)]
        O2_WFM_Data_Filtered=df2.reset_index(drop=True)
        O2_WFM_Data_Filtered['Statusreason'].drop_duplicates(keep='last', inplace=True)

        #O2_WFM_Data_Filtered.drop_duplicates(keep='last', inplace=True)

        Rev_Statusreason=O2_WFM_Data_Filtered['Statusreason'].iloc[::-1]
        #print 'Stauts Reason after droping duplicates', Rev_Statusreason
        #print 'Taken', Wo_Ref

        #print "Stauts Reasons taken for Iteration",Rev_Statusreason
        #print type(Rev_Statusreason)
        count_Nan=0
        #print '#############Yesterday_Comment#############', Yesterday_Comment
        Babcock=['3pp','mewp','pass to 3pp','3pp form','antenna','babcock','attached 3pp','mittie','mitie','mitre','attached 3pp request','3pp attached','3pp form attached','passing to 3pp', 'pass to babcock','passing to babcock', 'pass for riggers', 'passing for riggers','assign to 3pp', 'assign to babcock', 'assign to riggers', 'email sen to 3pp', 'email sent to babcock','cherry picker','cp req','cp need']
        Spares_ETA=['ordered','order','nbd','backorder','backlog','hws','eta','spare','spares','business day','part due','tomorrow','reschedule','part is coming from','need to order','part coming tommorrow','see il','spare ordered for delivery to bb','spare ordered','ordered spare']
        Spares_No_ETA=['ordered','order','nbd','faulty','hws','backlog','failure','spare','spares','awaiting eta','business day','part due','no update on eta','part is coming from','reschedule','tomorrow','arriving tomorrow','delivery tomorrow','part is not available','no eta','no eta as of now','bo ','back order','backorder','part not arrived','parts not arrived','part coming tommorrow','parts coming tommorrow','no update for spares','no update for spares eta','no updated on spare','spare resourcing issue','spares resourcing issue']
        Access_Requested=['access denied','rebook','no access','no response','awaiting response','awaiting conf','awaiting','mail sent to sp','mailed sp','email for sp','awaiting access confirmation','awaiting confirmation','need 48 hrs of notice','needs 48 hrs of notice','notice req','access refused','egi to arrange/confirm access','egi to arrange access','egi to confirm access','Access route is blocked','Access blocked','out of the office','further investigation required by climber','climber required','climb required','days notice','day notice','key','lock']
        Reschedule=['reschedule','access','denied','arrange access','high priority work','no response']
        Spares_Back_order=['backorder','part','order','ordered','part not in stock','unable','locate','hws','not in stock','bo']
        C3DTC=['awaiting access confirmation','access not confirmed','no response','passing to ctil','pass to ctil','email sent to ctil']
        Estates_Escalation1=['no responce','temporary','shut','access','block','padlock','passing to ctil','pass to ctil','email sent to ctil','ctil','key','lock','rent','gate']
        EOS=['eos','shift','time','end of shift','reschedule','further investigation','reassign','key','oot','cop','no time','not going','hpw','hp','shift has completed','out of time','no time to deal','not enough time','insufficient','finish time']
        HP=['hp job','busy on','busy with','P2','P3','p2','p3','prevent me','divert','diverted','cant attend','hp','ticket','previous','eos','shift','no time','no enough time','not enough time']
        Resource_Skill_Level=['climber','ret','climbing team','pim','interference','pim test','site master','climb','two man','2nd man','2 man','experienced','capability','support','swap']
        Res_Tools=['climber','climbing team','ret','pim','interference','pim test','site master','sitemaster','climb','two man','two men','2nd man','2 man','ladder','test','tested','permanent','dummy','load','health','safety']
        Reschedule1=['reschedule','pass for tomorrow','assign','reallocate','daylight']
        Support_Customer1=['script','o2','ats','comms','file','power','service desk','integration','sd','optimisation','config','unable','mail','nigel','commissi','commissioning','waiting call back from nmc','huawei router','mitie','3pp','bt fault','bt','theft','crime','waiting']
        JM_OLO=['joint meet','jm','bt','nte','jointmeet','ats']
        Access_Site_Inaccessible=['no answer','daylight','closed','shut','closed','dark','no work','lock','not answered','no access','access','obas','reschedule','denied','2man','2 man','two man','cannot access','ladder']
        Access_Req_Booking=['no access','access','key','reallocate']
        H_and_S=['not safe','safe','night','climb','daylight','tower','H&S','dark','ladder','rain','ice','weather','snow','health','safety']
        Furtherinfo=['further info','pim','pimtester','pim tester','climb','antenna','cow','crq','ats','service desk']
        Climate=['rain','dark','weather','snow','ice','wind']
        Additional_Effort=['revisit','climb','pim','pim tester','pim test','site master','sitemaster','climb','antenna','2man','2nd man','two man']
        pat_given=re.compile(r'[Gg]iven')
        

        def Babcock_Action_Required(Yesterday_Comment_Rev,y):
            for k in Babcock:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
        
        def Reschedule(Yesterday_Comment_Rev,y):
            for k in Reschedule1:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Resource_SkillLevel(Yesterday_Comment_Rev,y):
            for k in Resource_Skill_Level:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def C3_DTC(Yesterday_Comment_Rev,y):
            for k in C3DTC:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Estates_Escalation(Yesterday_Comment_Rev,y):
            for k in Estates_Escalation1:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
                    
        def Spares_ETA_Given(Yesterday_Comment_Rev,Spares_ETA,y):
            #print 'Entered Spares_ETA_Given Function'
            for k in Spares_ETA:
                for x in Yesterday_Comment_Rev:
                #    print 'Taken log entry: ',x ,'and spares list as: ',k
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return

        def Spares_No_ETA_Given(Yesterday_Comment_Rev,y):
            for k in Spares_No_ETA:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Spares_Backorder(Yesterday_Comment_Rev,y):
            for k in Spares_Back_order:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        #print k
                        return
                  
        def Access_Req(Yesterday_Comment_Rev,y):
            for k in Access_Requested:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Access_Req_Book(Yesterday_Comment_Rev,y):
            for k in Access_Req_Booking:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Support_Customer(Yesterday_Comment_Rev,y):
            for k in Support_Customer1:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                
        def Resource_Tools(Yesterday_Comment_Rev,y):
            for k in Res_Tools:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
        def Support_JM_OLO(Yesterday_Comment_Rev,y):
            for k in JM_OLO:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
        def HP_Job(Yesterday_Comment_Rev,y):
            for k in HP:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def End_of_Shift(Yesterday_Comment_Rev,y):
            for k in EOS:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Site_Inaccessible(Yesterday_Comment_Rev,y):
            for k in Access_Site_Inaccessible:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Health_Safety(Yesterday_Comment_Rev,y):
            for k in H_and_S:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
        
        def Further_Info(Yesterday_Comment_Rev,y):
            for k in Furtherinfo:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Climatology(Yesterday_Comment_Rev,y):
            for k in Climate:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
        def Additional_Eff(Yesterday_Comment_Rev,y):
            for k in Additional_Effort:
                for x in Yesterday_Comment_Rev:
                    if k in x:
                        ws.cell(row=i, column=26).value=k
                        ws.cell(row=i, column=27).value=y
                        ws.cell(row=i, column=28).value=Yesterday_Comment_Rev.index(x)+1
                        ws.cell(row=i, column=29).value=x.strip()
                        return
                    
                
                
        if len(Rev_Statusreason)==Rev_Statusreason.isnull().sum():
            ws.cell(row=i, column=26).value='Blank Status Reason'
            ws.cell(row=i, column=27).value='Blank Status Reason'
            ws.cell(row=i, column=28).value='Blank Status Reason'
            ws.cell(row=i, column=29).value='Blank Status Reason'
        else:
            #print 'Picked after droping Nan', Rev_Statusreason.dropna(axis=0, how='any')
            Rev_Statusreason1=Rev_Statusreason.dropna(axis=0, how='any')
            for y in Rev_Statusreason1:
                if ('3PP' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Babcock_Action_Required(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Spares' in y) and len(re.findall(pat_given,y))==1:
                    print 'Taken', y , 'for ',Wo_Ref
                    Spares_ETA_Given(Yesterday_Comment_Rev,Spares_ETA,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Spares' in y) and ('No ETA' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Spares_No_ETA_Given(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Spares' in y) and ('Backorder' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Spares_Backorder(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('C3' in y) and ('DTC' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    C3_DTC(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Resource' in y) and ('Skill' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Resource_SkillLevel(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Reschedule' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Reschedule(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                #elif ('Access' in y) and ('Req' in y):
                #    print 'Taken', y , 'for ',Wo_Ref
                #   Access_Req(Yesterday_Comment_Rev,y)
                #    if ws.cell(row=i, column=26).value!=None:
                #        break
                elif ('Support' in y) and ('Customer' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Support_Customer(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif (('Resource' in y) and ('Tools' in y)) or (('Test' in y) and ('Equipment' in y)):
                    print 'Taken', y , 'for ',Wo_Ref
                    Resource_Tools(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('OLO' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Support_JM_OLO(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Other' in y) and ('High' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    HP_Job(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('End' in y) and ('Shift' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    End_of_Shift(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Site' in y) and ('Inaccessible' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Site_Inaccessible(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Health' in y) and ('Safety' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Health_Safety(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Estates' in y) and ('Escalation' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Estates_Escalation(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Further' in y) and ('Info' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Further_Info(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Climatology' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Climatology(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Requires' in y) and ('Booking' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Access_Req_Book(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                elif ('Additional' in y) and ('Effort' in y):
                    print 'Taken', y , 'for ',Wo_Ref
                    Additional_Eff(Yesterday_Comment_Rev,y)
                    if ws.cell(row=i, column=26).value!=None:
                        break
                
                else:
                    ws.cell(row=i, column=26).value='No Matching function for this AH'
                    ws.cell(row=i, column=27).value='No Matching function for this AH'
                    ws.cell(row=i, column=28).value='No Matching function for this AH'
                    ws.cell(row=i, column=29).value='No Matching function for this AH'
        
        if ws.cell(row=i, column=26).value==None:
            ws.cell(row=i, column=26).value='No enough information for re-work'
            ws.cell(row=i, column=27).value='No enough information for re-work'
            ws.cell(row=i, column=28).value='No enough information for re-work'
                

    wb.save('C:\Users\eubcefm\PycharmProjects\Re-Work\Corrective_Task.xlsx')

print '*************Done with Script*************'
